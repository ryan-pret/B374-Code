from openpyxl import *
# load excel with its path
wrkbk = load_workbook("C:\\work\\SeleniumPython.xlsx")
# to get the active work sheet
sh = wrkbk.active
# to print the maximum number of occupied rows in console
print(sh.max_row)
# to print the maximum number of occupied columns in console
print(sh.max_column)
# to get all the values from the excel and traverse through the rows
for r in range(1,max_row+1):
# to check the value in column 1
    if(sh.cell(row=r, column=1).value == "2":
# to traverse through the columns
        for c in range(2,max_column+1):
# to get all the values
            print(sh.cell(row=r, column=c).value)